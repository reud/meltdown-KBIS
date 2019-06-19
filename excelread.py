# -*- coding: utf-8 -*-
import traceback
import openpyxl
import json


class Member(object):
    """
    メンバーの情報をまとめるクラス
    """

    def __init__(self, name: str, gen: int, money: int, id: str = '-1', bikou: str = 'なし'):
        self.name = name
        self.gen = gen
        self.money = money
        self.id = id
        self.bikou = bikou

    def printData(self):
        """
        メンバの情報をLINENotifyで出力する。
        :return:
        """
        print(f'{self.name}さん {self.gen}G {self.money}円 id is {self.id} 備考:{self.bikou} ')

    def output_data(self) -> str:
        """
        メンバの情報を文字列にして返す
        :return　str: メンバの情報文字列
        """
        return f'{self.name}さん {self.gen}G {self.money}円 id is {self.id} 備考:{self.bikou} '

    def get_dict(self) -> dict:
        return {"name": self.name, "gen": self.gen, "money": self.money, "id": self.id, "bikou": self.bikou}


class Group(object):
    """
    班の情報を格納するクラス
    班には予備費も含まれる
    """

    def __init__(self, name: str, money: int):
        self.name = name
        self.money = money

    def printData(self):
        """
        班の情報をLINENotifyで出力するクラス
        :return None:
        """
        print(f'{self.name}: 残高{self.money}')


class Manager(object):
    """
    ユーザのデータを管理するクラス
    班の情報も格納される
    """

    def __init__(self, path: str):
        """
        エクセルファイルを読み込んで、メンバ変数に保管します。
        :param path: エクセルファイルの置いてある場所
        """
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        yosan_sheet = workbook['予算']
        self.memberlist: list = []
        self.path = path
        FROM = 15
        TO = 21

        for i in range(FROM, TO + 1):
            ws = workbook[str(i) + 'G']
            for mem in range(4, 30):
                hisname = ws['B' + str(mem)].value
                if (hisname):
                    bikou = ws['E' + str(mem)].value
                    hismoney = ws['C' + str(mem)].value
                    self.memberlist.append(Member(name=hisname, gen=i, money=hismoney if hismoney else 0,
                                                  bikou=bikou if bikou else 'なし'))

        self.groups = []
        self.groups.append(Group('全体残高', yosan_sheet['D3'].value))

        self.groups.append(Group('設計班', yosan_sheet['L9'].value))
        self.groups.append(Group('翼班', yosan_sheet['L10'].value))
        self.groups.append(Group('コクピ班', yosan_sheet['L11'].value))
        self.groups.append(Group('接合班', yosan_sheet['L12'].value))
        self.groups.append(Group('電装班', yosan_sheet['L13'].value))
        self.groups.append(Group('デザイン班', yosan_sheet['L14'].value))
        self.groups.append(Group('予備費', yosan_sheet['L15'].value))
        # self.file_import()

    def setId(self, name: str, id: str, strict: bool = False) -> str:
        """
        メンバーリストを更新する時に使用する関数
        :param name: 本名
        :param id: user_id 初期値(何も入っていない時)は-1
        :param strict:
            Trueの場合はもともとidが入っていても問答無用で更新
            Falseの場合はもともとidが入っていたらKeyError
        :return:
            変更結果を文字列にして返す。
        """
        for i in range(len(self.memberlist)):
            if (self.memberlist[i].name == name):
                if (self.memberlist[i].id == "-1"):
                    self.memberlist[i].id = id
                    return f'{name} new id->{self.memberlist[i].id}'
                else:
                    if (not strict):
                        raise KeyError('idが初期値でないメンバのidを更新しようとしました。')
                    else:
                        strings = '-strict set happen-'
                        self.memberlist[i].id = id
                        return strings + f'\n{name} new id->{self.memberlist[i].id}'

        raise NameError(f'{name} not found \n その名前のメンバは見つかりませんでした。 searched from {self.memberlist}')

    def getFromId(self, id: str) -> Member:
        """
        user_idからメンバーを検索する。
        見つからなかった場合はKeyError
        :param id: user_id
        :return Member: Memberオブジェクト
        """
        for i in self.memberlist:
            if (i.id == id):
                return i
        raise KeyError('そのidのメンバは見つかりませんでした。')

    def getFromName(self, name: str) -> Member:
        """
        user_idからメンバーを検索する。
        見つからなかった場合はKeyError
        :param name: 本名
        :return: Memberオブジェクト
        """
        for i in self.memberlist:
            if (i.name == name):
                return i
        raise KeyError('名前が見つかりませんでした')

    def printData(self):
        """
        LINENotifyに全ユーザのデータを出力する
        :return: なし
        """
        for i in self.memberlist:
            i.printData()

    def output_data(self) -> str:
        """
        全ユーザデータを文字列として返す
        printData()と内容は一緒
        :return: 全ユーザのデータを文字列にしたもの
        """
        strings = ''
        for i in self.memberlist:
            strings += i.output_data() + '\n'
        return strings

    def outputIdConnection(self) -> str:
        """
        全ユーザの名前とIDを文字列にして返す
        :return:  全ユーザの名前データとIDデータ
        """
        out = ''
        for i in self.memberlist:
            out += f'{i.name},{i.id}\n'
        return out

    def file_import(self):
        """
        excel-id.txtファイルを読み込んでIDと名前の結び付けを行う関数
        :return: なし
        """
        try:
            file = open('excel-id.txt')
            string_list = file.readlines()
            string_list = [i.strip() for i in string_list]
            list_in_list = []
            for i in string_list:
                list_in_list.append(i.split(','))
            print(list_in_list)
            for i in list_in_list:
                if len(i) < 2:
                    raise IndexError(f'要素が1以下の配列にアクセスしようとしています。　内容: {list_in_list}')
                self.setId(name=i[0], id=i[1], strict=True)
                pass

        except:
            print(traceback.format_exc())

    def get_json(self) -> str:
        return json.dumps({"members": [m.get_dict() for m in self.memberlist] }, ensure_ascii=False).encode("utf-8")
